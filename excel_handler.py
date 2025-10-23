"""
Модуль для работы с Excel файлами (XLSX) в Yo Store
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Optional
import json
from datetime import datetime
import io

class ExcelHandler:
    """Класс для работы с Excel файлами"""
    
    def __init__(self):
        self.product_columns = [
            'sku', 'name', 'description', 'level0', 'level1', 'level2', 'brand', 
            'price', 'currency', 'stock', 'image_url', 'specifications'
        ]
        
        self.price_columns = [
            'sku', 'name', 'price', 'old_price', 'currency'
        ]
    
    def create_products_template(self) -> bytes:
        """Создать шаблон Excel файла для добавления товаров"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        # Заголовки
        headers = [
            'SKU товара', 'Название товара*', 'Описание', 'Основная категория (level0)*', 'Подкатегория (level1)*', 'Детальная категория (level2)*', 'Бренд',
            'Цена*', 'Валюта', 'Количество на складе', 'URL изображения (через запятую)', 'Характеристики (JSON)'
        ]
        
        # Добавить заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавить примеры данных
        examples = [
            ['IPHONE16Pro-256GB-TitaniumNatural', 'iPhone 16 Pro 256GB Titanium Natural', 'Новейший iPhone с титановым корпусом', 'Смартфоны', '16 Series', '16 Pro', 'Apple', 
             89990, 'RUB', 10, '/static/images/products/iphone16pro/titanium-natural/1.jpg, /static/images/products/iphone16pro/titanium-natural/2.jpg', '{"color": "Titanium Natural", "disk": "256GB", "sim_config": "Dual SIM"}'],
            ['MACBOOKAIRM2-512GB-Silver', 'MacBook Air M2 512GB Silver', 'Легкий и мощный ноутбук', 'Ноутбуки', 'Air Series', 'Air M2', 'Apple', 
             149990, 'RUB', 5, '/static/images/products/macbookairm2/silver/1.jpg, /static/images/products/macbookairm2/silver/2.jpg', '{"color": "Silver", "ram": "16GB", "disk": "512GB"}']
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Добавить лист с изображениями
        images_ws = wb.create_sheet("Изображения")
        images_headers = [
            'Модель (level_2)*', 'Цвет*', 'URL изображений (через запятую)*'
        ]
        
        # Добавить заголовки для изображений
        for col, header in enumerate(images_headers, 1):
            cell = images_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавить примеры изображений
        images_examples = [
            ['iPhone 16 Pro', 'Titanium Natural', '/static/images/products/iphone16pro/titanium-natural/1.jpg, /static/images/products/iphone16pro/titanium-natural/2.jpg, /static/images/products/iphone16pro/titanium-natural/3.jpg'],
            ['iPhone 16 Pro', 'Titanium Black', '/static/images/products/iphone16pro/titanium-black/1.jpg, /static/images/products/iphone16pro/titanium-black/2.jpg, /static/images/products/iphone16pro/titanium-black/3.jpg'],
            ['MacBook Air M2', 'Silver', '/static/images/products/macbookairm2/silver/1.jpg, /static/images/products/macbookairm2/silver/2.jpg, /static/images/products/macbookairm2/silver/3.jpg'],
            ['MacBook Air M2', 'Space Gray', '/static/images/products/macbookairm2/space-gray/1.jpg, /static/images/products/macbookairm2/space-gray/2.jpg, /static/images/products/macbookairm2/space-gray/3.jpg']
        ]
        
        for row, example in enumerate(images_examples, 2):
            for col, value in enumerate(example, 1):
                images_ws.cell(row=row, column=col, value=value)
        
        # Автоподбор ширины колонок для изображений
        for column in images_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            images_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавить лист с категориями
        categories_ws = wb.create_sheet("Категории")
        categories_ws.append(['ID', 'Название', 'Описание', 'Иконка'])
        
        categories_data = [
            [1, 'Телефоны', 'Смартфоны и мобильные телефоны', '📱'],
            [2, 'Ноутбуки', 'Портативные компьютеры', '💻'],
            [3, 'Игровые приставки', 'Игровые консоли', '🎮'],
            [4, 'Наушники', 'Аудио устройства', '🎧'],
            [5, 'Планшеты', 'Планшетные компьютеры', '📱'],
            [6, 'Умные колонки', 'Голосовые помощники', '🔊'],
            [7, 'Телевизоры', 'Телевизионные панели', '📺'],
            [8, 'Умные часы', 'Носимые устройства', '⌚']
        ]
        
        for category in categories_data:
            categories_ws.append(category)
        
        # Стилизация листа категорий
        for col in range(1, 5):
            cell = categories_ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Автоподбор ширины колонок
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранить в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def create_prices_template(self) -> bytes:
        """Создать шаблон Excel файла для обновления цен"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Цены"
        
        # Заголовки
        headers = [
            'SKU товара*', 'Название товара', 'Новая цена*', 'Старая цена', 'Валюта'
        ]
        
        # Добавить заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавить примеры данных
        examples = [
            ['IPHONE16Pro-256GB-TitaniumNatural', 'iPhone 16 Pro 256GB Titanium Natural', 89990, 99990, 'RUB'],
            ['MACBOOKProM3-512GB-Silver', 'MacBook Pro M3 512GB Silver', 199990, 219990, 'RUB']
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранить в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def parse_products_excel(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Парсить Excel файл с товарами"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(io.BytesIO(file_content), sheet_name='Товары')
            
            # Проверяем обязательные колонки
            required_columns = ['Название товара*', 'Основная категория (level0)*', 'Подкатегория (level1)*', 'Детальная категория (level2)*', 'Цена*']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}")
            
            products = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Пропускаем пустые строки (проверяем обязательные поля)
                    if (pd.isna(row['Название товара*']) or 
                        pd.isna(row['Основная категория (level0)*']) or 
                        pd.isna(row['Подкатегория (level1)*']) or 
                        pd.isna(row['Детальная категория (level2)*']) or 
                        pd.isna(row['Цена*'])):
                        continue
                    
                    # Парсим характеристики
                    specifications = {}
                    if not pd.isna(row.get('Характеристики (JSON)', '')):
                        try:
                            specifications = json.loads(str(row['Характеристики (JSON)']))
                        except json.JSONDecodeError:
                            specifications = {}
                    
                    # Извлекаем специфичные поля из specifications или из отдельных колонок
                    color = specifications.get('color', '') if specifications else ''
                    disk = specifications.get('disk', specifications.get('memory', '')) if specifications else ''
                    ram = specifications.get('ram', '') if specifications else ''
                    sim_config = specifications.get('sim_config', specifications.get('sim_type', '')) if specifications else ''
                    
                    product = {
                        'sku': str(row.get('SKU товара', '')).strip() if not pd.isna(row.get('SKU товара', '')) else '',
                        'name': str(row['Название товара*']).strip(),
                        'description': str(row.get('Описание', '')).strip() if not pd.isna(row.get('Описание', '')) else '',
                        'level0': str(row['Основная категория (level0)*']).strip(),
                        'level1': str(row['Подкатегория (level1)*']).strip(),
                        'level2': str(row['Детальная категория (level2)*']).strip(),
                        'brand': str(row.get('Бренд', '')).strip() if not pd.isna(row.get('Бренд', '')) else '',
                        'price': float(row['Цена*']),
                        'currency': str(row.get('Валюта', 'RUB')).strip().upper() if not pd.isna(row.get('Валюта', 'RUB')) else 'RUB',
                        'stock': int(row.get('Количество на складе', 0)) if not pd.isna(row.get('Количество на складе', 0)) else 0,
                        'image_url': str(row.get('URL изображения (через запятую)', '')).strip() if not pd.isna(row.get('URL изображения (через запятую)', '')) else '',
                        'specifications': specifications,
                        'color': color,
                        'disk': disk,
                        'ram': ram,
                        'sim_config': sim_config
                    }
                    
                    products.append(product)
                    
                except Exception as e:
                    errors.append(f"Строка {index + 2}: {str(e)}")
            
            if errors:
                raise ValueError(f"Ошибки при парсинге: {'; '.join(errors)}")
            
            return products
            
        except Exception as e:
            raise ValueError(f"Ошибка при чтении Excel файла: {str(e)}")
    
    def parse_prices_excel(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Парсить Excel файл с ценами"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(io.BytesIO(file_content), sheet_name='Цены')
            
            # Проверяем обязательные колонки
            required_columns = ['SKU товара*', 'Новая цена*']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}")
            
            prices = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Пропускаем пустые строки
                    if pd.isna(row['SKU товара*']) or pd.isna(row['Новая цена*']):
                        continue
                    
                    price_data = {
                        'sku': str(row['SKU товара*']).strip(),
                        'name': str(row.get('Название товара', '')).strip(),
                        'price': float(row['Новая цена*']),
                        'old_price': float(row.get('Старая цена', row['Новая цена*'])),
                        'currency': str(row.get('Валюта', 'RUB')).strip().upper()
                    }
                    
                    prices.append(price_data)
                    
                except Exception as e:
                    errors.append(f"Строка {index + 2}: {str(e)}")
            
            if errors:
                raise ValueError(f"Ошибки при парсинге: {'; '.join(errors)}")
            
            return prices
            
        except Exception as e:
            raise ValueError(f"Ошибка при чтении Excel файла: {str(e)}")
    
    def parse_images_excel(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Парсить Excel файл с изображениями"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(io.BytesIO(file_content), sheet_name='Изображения')
            
            # Проверяем обязательные колонки
            required_columns = ['Модель (level_2)*', 'Цвет*', 'URL изображений (через запятую)*']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}")
            
            images = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Пропускаем пустые строки
                    if pd.isna(row['Модель (level_2)*']) or pd.isna(row['Цвет*']) or pd.isna(row['URL изображений (через запятую)*']):
                        continue
                    
                    # Парсим URL изображений
                    image_urls_str = str(row['URL изображений (через запятую)*']).strip()
                    if not image_urls_str:
                        continue
                    
                    # Разделяем URL по запятой и очищаем от пробелов
                    image_urls = [url.strip() for url in image_urls_str.split(',') if url.strip()]
                    
                    if not image_urls:
                        continue
                    
                    image_data = {
                        'level_2': str(row['Модель (level_2)*']).strip(),
                        'color': str(row['Цвет*']).strip(),
                        'img_list': image_urls
                    }
                    
                    images.append(image_data)
                    
                except Exception as e:
                    errors.append(f"Строка {index + 2}: {str(e)}")
            
            if errors:
                raise ValueError(f"Ошибки при парсинге: {'; '.join(errors)}")
            
            return images
            
        except Exception as e:
            raise ValueError(f"Ошибка при чтении Excel файла: {str(e)}")
    
    def export_products_to_excel(self, products: List[Dict[str, Any]]) -> bytes:
        """Экспортировать товары в Excel файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        # Заголовки
        headers = [
            'ID', 'SKU', 'Название', 'Категория', 'Уровень 0', 'Уровень 1', 'Уровень 2', 'Бренд',
            'Цена', 'Старая цена', 'Валюта', 'Скидка %', 'На складе', 'Доступен', 'Изображения', 'Кол-во изображений', 'Дата создания'
        ]
        
        # Добавить заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавить данные
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product.get('id'))
            ws.cell(row=row, column=2, value=product.get('sku'))
            ws.cell(row=row, column=3, value=product.get('name'))
            ws.cell(row=row, column=4, value=product.get('category_name'))
            ws.cell(row=row, column=5, value=product.get('level0'))
            ws.cell(row=row, column=6, value=product.get('level1'))
            ws.cell(row=row, column=7, value=product.get('level2'))
            ws.cell(row=row, column=8, value=product.get('brand'))
            ws.cell(row=row, column=9, value=product.get('price'))
            ws.cell(row=row, column=10, value=product.get('old_price'))
            ws.cell(row=row, column=11, value=product.get('currency'))
            ws.cell(row=row, column=12, value=product.get('discount_percentage'))
            ws.cell(row=row, column=13, value=product.get('stock'))
            ws.cell(row=row, column=14, value=product.get('is_available'))
            ws.cell(row=row, column=15, value=product.get('image_text'))
            ws.cell(row=row, column=16, value=product.get('images_count'))
            ws.cell(row=row, column=17, value=product.get('created_at'))
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранить в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
