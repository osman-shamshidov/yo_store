from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Request, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, RedirectResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from database import get_db
from models import Product, Category, CurrentPrice, ProductImage, Level2Description
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import json
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from excel_handler import ExcelHandler
from manual_price_manager import manual_price_manager
from config import Config
import os

def normalize_model_key(model_key: str) -> str:
    """Нормализовать ключ модели для поиска в файловой системе"""
    # Приводим к нижнему регистру
    model_key = model_key.lower()

    
    # Если не найдено совпадение, возвращаем как есть
    return model_key

def normalize_color_name(color: str) -> str:
    """Нормализовать название цвета для поиска в файловой системе"""
    # Просто приводим к нижнему регистру и заменяем пробелы на дефисы
    return color.lower().replace(' ', '-')

def get_product_images(product, db: Session):
    """Получить массив изображений товара из таблицы ProductImage"""
    images = []
    
    # Сначала попробуем получить изображения из specifications
    try:
        specs = json.loads(product.specifications) if product.specifications else {}
        images_data = specs.get('images', [])
        
        for img_data in images_data:
            if isinstance(img_data, dict):
                # Новый формат: {"url": "...", "alt": "..."}
                images.append(img_data["url"])
            elif isinstance(img_data, str):
                # Новый формат: массив строк
                images.append(img_data)
                
    except json.JSONDecodeError:
        pass
    
    # Если не нашли в specifications, ищем в таблице ProductImage по (level_2, color)
    if not images and product.level_2 and product.color:
        try:
            product_image = db.query(ProductImage).filter(
                ProductImage.level_2 == product.level_2,
                ProductImage.color == product.color
            ).first()
            
            if product_image and product_image.img_list:
                images_data = json.loads(product_image.img_list)
                
                # Обработка double-encoded JSON (если после парсинга получили строку)
                if isinstance(images_data, str):
                    images_data = json.loads(images_data)
                
                # Теперь обрабатываем массив
                if isinstance(images_data, list):
                    for img_data in images_data:
                        if isinstance(img_data, dict):
                            images.append(img_data["url"])
                        elif isinstance(img_data, str):
                            images.append(img_data)
        except (json.JSONDecodeError, TypeError):
            pass
    
    # Если изображений нет, вернем пустой список
    return images

def parse_images_from_string(images_str: str) -> List[str]:
    """Парсить строку изображений разделенных запятыми в JSON массив"""
    if not images_str or not images_str.strip():
        return []
    
    # Разделяем по запятой и очищаем от пробелов
    image_urls = [url.strip() for url in str(images_str).split(',') if url.strip()]
    
    # Конвертируем в JSON массив строк
    return json.dumps(image_urls)

app = FastAPI(title="Yo Store API", version="1.0.0")

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# --- Simple Admin Auth (cookie-based) ---
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'yo_admin')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'yo_admin')
ADMIN_SESSION_TOKEN = "yo_admin_session_token_v1"

def is_admin_authenticated(request: Request) -> bool:
    token = request.cookies.get("admin_session")
    return token == ADMIN_SESSION_TOKEN

def require_admin(request: Request):
    if not is_admin_authenticated(request):
        raise HTTPException(status_code=401, detail="Unauthorized")
    return True

# Pydantic models for API
class ProductResponse(BaseModel):
    id: int
    sku: Optional[str] = None  # SKU только у вариантов, для общих карточек = None
    name: str
    description: str = ""  # По умолчанию пустая строка, а не None
    brand: str
    model: str
    category_name: str
    level_2: Optional[str] = None  # Название группы товаров (например "iPhone 16 Pro Max")
    image_url: str
    images: List[str] = []  # Массив изображений
    specifications: dict
    price: Optional[float] = 0.0
    old_price: Optional[float] = 0.0
    discount_percentage: Optional[float] = 0.0
    currency: str = "RUB"
    is_available: bool = True
    
    class Config:
        from_attributes = True

class CategoryResponse(BaseModel):
    id: int
    name: str
    description: str
    icon: str
    product_count: int
    parent_category_id: Optional[int] = None
    brand: Optional[str] = None
    is_subcategory: bool = False
    subcategories: List['CategoryResponse'] = []
    
    class Config:
        from_attributes = True

class ProductDetailResponse(BaseModel):
    id: int
    sku: str  # Уникальный SKU товара
    name: str
    description: str
    brand: str
    model: str
    category_name: str
    image_url: str
    images: List[str] = []  # Массив изображений
    specifications: dict
    price: float
    old_price: float
    discount_percentage: float
    currency: str
    is_available: bool
    created_at: str
    
    class Config:
        from_attributes = True

# --- Helpers ---
def ensure_category_exists(db: Session, level0: Optional[str], level1: Optional[str] = None, level2: Optional[str] = None) -> None:
    """Создать записи в таблице Category для уровней, если их нет."""
    try:
        # level_0 only
        if level0:
            exists_l0 = db.query(Category).filter(Category.level_0 == level0, Category.level_1 == None, Category.level_2 == None).first()
            if not exists_l0:
                db.add(Category(level_0=level0, level_1=None, level_2=None))
                db.flush()
        # level_1
        if level0 and level1:
            exists_l1 = db.query(Category).filter(Category.level_0 == level0, Category.level_1 == level1, Category.level_2 == None).first()
            if not exists_l1:
                db.add(Category(level_0=level0, level_1=level1, level_2=None))
                db.flush()
        # level_2
        if level0 and level1 and level2:
            exists_l2 = db.query(Category).filter(Category.level_0 == level0, Category.level_1 == level1, Category.level_2 == level2).first()
            if not exists_l2:
                db.add(Category(level_0=level0, level_1=level1, level_2=level2))
                db.flush()
    except Exception:
        # Не прерываем основной процесс импорта из-за проблем с категориями
        pass

# API Routes
@app.get("/")
async def root():
    return {"message": "Yo Store API", "version": "1.0.0"}

@app.get("/test-products")
async def test_products(db: Session = Depends(get_db)):
    """Простой тестовый endpoint для проверки"""
    products = db.query(Product).limit(2).all()
    return [{"id": p.id, "sku": p.sku, "name": p.name, "level_0": p.level_0} for p in products]

@app.get("/categories")
async def get_categories(db: Session = Depends(get_db)):
    """Get all categories grouped by level_0"""
    # Получить уникальные категории из таблицы Category с GROUP BY
    from sqlalchemy import func
    categories = db.query(
        Category.level_0,
        func.max(Category.description).label('description'),
        func.max(Category.icon).label('icon')
    ).filter(
        Category.level_0.isnot(None)
    ).group_by(Category.level_0).all()
    
    result = []
    for level_0, description, icon in categories:
        # Подсчитать товары в этой категории
        product_count = db.query(Product).filter(Product.level_0 == level_0).count()
        
        result.append({
            "id": abs(hash(level_0)) % 1000000,  # Генерируем положительный ID из хэша
            "name": level_0,
            "description": description or f"Категория {level_0}",
            "icon": icon or "📦",
            "product_count": product_count,
            "level_0": level_0
        })
    
    # Сортируем по количеству товаров в убывающем порядке
    result.sort(key=lambda x: x["product_count"], reverse=True)
    
    return result

@app.get("/all-products", response_model=List[ProductResponse])
async def get_all_products(db: Session = Depends(get_db)):
    """Endpoint для получения всех товаров без группировки"""
    try:
        # Простой запрос всех товаров
        results = db.query(Product, CurrentPrice).outerjoin(
            CurrentPrice, Product.sku == CurrentPrice.sku
        ).order_by(Product.level_0, Product.level_1, Product.level_2, Product.sku).all()
        
        print(f"📊 Найдено {len(results)} результатов в БД")
        
        products = []
        for idx, result in enumerate(results):
            print(f"🔄 Обрабатываем товар {idx + 1}/{len(results)}: ID {result[0].id}")
            product = result[0]  # Product
            price = result[1]    # CurrentPrice или None
            
            # Получаем данные о цене с безопасными значениями по умолчанию
            if price is None:
                product_price = 0.0
                product_old_price = 0.0
                product_discount = 0.0
                product_currency = "RUB"
            else:
                product_price = price.price if price.price is not None else 0.0
                product_old_price = price.old_price if price.old_price is not None else 0.0
                product_discount = price.discount_percentage if price.discount_percentage is not None else 0.0
                product_currency = price.currency if price.currency else "RUB"
            
            # Получаем изображения
            images = get_product_images(product, db)
            image_url = images[0] if images else "/static/images/placeholder.jpg"
            
            # Получаем спецификации
            try:
                specifications = json.loads(product.specifications) if product.specifications else {}
            except json.JSONDecodeError:
                specifications = {}
            
            # Проверяем description
            desc = ""
            if True:  # description всегда None, так как поле удалено
                print(f"⚠️  Товар ID {product.id} имеет description=None, заменяем на пустую строку")
            
            products.append(ProductResponse(
                id=product.id,
                sku=product.sku,
                name=product.name,
                description=desc,
                brand=product.brand,
                model=product.level_2 or "",
                category_name=f"{product.level_0} / {product.level_1} / {product.level_2}" if product.level_1 and product.level_2 else product.level_0,
                level_2=product.level_2,
                image_url=image_url,
                images=images,
                specifications=specifications,
                price=product_price,
                old_price=product_old_price,
                discount_percentage=product_discount,
                currency=product_currency,
                is_available=True
            ))
        
        return products
    except Exception as e:
        print(f"❌ Ошибка в get_all_products: {e}")
        return []

@app.get("/products", response_model=List[ProductResponse])
async def get_products(
    brand: Optional[str] = None,
    level0: Optional[str] = None,
    level1: Optional[str] = None,
    level2: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """Get unique product models (grouped by level2) with optional hierarchical filters"""
    # Простая логика: получаем все товары, затем группируем в Python
    query = db.query(Product, CurrentPrice).outerjoin(
        CurrentPrice, Product.sku == CurrentPrice.sku
    )
    
    # Применяем фильтры
    filters = []
    
    if brand:
        filters.append(Product.brand == brand)
    if level0:
        filters.append(Product.level_0 == level0)
    if level1:
        filters.append(Product.level_1 == level1)
    if level2:
        filters.append(Product.level_2 == level2)
        
    
    # Применяем фильтры
    if filters:
        query = query.filter(and_(*filters))
    
    # Используем подкеру для получения одного представительного товара из каждой модели
    subquery = db.query(
        func.min(Product.id).label('id')
    ).filter(*filters).group_by(Product.level_2, Product.brand).subquery()
    
    # Теперь получаем только те товары, которые являются представителями групп
    final_query = db.query(Product, CurrentPrice).outerjoin(
        CurrentPrice, Product.sku == CurrentPrice.sku
    ).outerjoin(subquery, Product.id == subquery.c.id).filter(
        subquery.c.id.isnot(None)
    ).order_by(Product.level_2, Product.id)
    
    # Применяем лимит и отступ
    results = final_query.offset(offset).limit(limit).all()
    
    products = []
    for result in results:
        product = result[0]  # Product


        price = result[1]    # CurrentPrice или None
        
        # Если цена отсутствует, используем значения по умолчанию
        if price is None:
            price_obj = type('Price', (), {
                'price': 0.0,
                'old_price': 0.0,
                'discount_percentage': 0.0,
                'currency': 'RUB'
            })()
        else:
            price_obj = price
        
        try:
            specifications = json.loads(product.specifications) if product.specifications else {}
        except json.JSONDecodeError:
            specifications = {}
        
        # Получаем описание из level2_descriptions
        desc = ""
        level2_specs = {}
        if product.level_2:
            level2_desc = db.query(Level2Description).filter(Level2Description.level_2 == product.level_2).first()
            if level2_desc:
                desc = level2_desc.description or ""
                if level2_desc.details:
                    try:
                        level2_specs = json.loads(level2_desc.details) if isinstance(level2_desc.details, str) else level2_desc.details
                    except json.JSONDecodeError:
                        level2_specs = {}
        
        # Объединяем характеристики из level2_descriptions с существующими specifications
        all_specifications = {**level2_specs, **specifications}
        
        images = get_product_images(product, db)
        
        # Получаем название категории из level полей
        category_name = product.level_0 or "Без категории"
        if product.level_1:
            category_name += f" / {product.level_1}"
        if product.level_2:
            category_name += f" / {product.level_2}"
        
        products.append(ProductResponse(
            id=product.id,
            sku=product.sku,
            name=product.name,
            description=desc,
            brand=product.brand,
            model=product.level_2 or "",
            category_name=category_name,
            level_2=product.level_2,
            image_url=images[0] if images else '',
            images=images,
            specifications=specifications,
            price=price_obj.price,
            old_price=price_obj.old_price,
            discount_percentage=price_obj.discount_percentage,
            currency=price_obj.currency,
        ))
    
    return products

@app.get("/products/{model}/variants")
async def get_model_variants(model: str, db: Session = Depends(get_db)):
    """Get all variants and their prices for a specific model (level_2)"""
    import urllib.parse
    # Декодируем URL параметр
    model = urllib.parse.unquote(model)
    
    # Сначала попробуем найти основной продукт с полными спецификациями
    main_product = db.query(Product).filter(
        Product.level_2 == model,
        Product.specifications.like('%variants%')
    ).first()
    
    if main_product:
        # Найден основной продукт с вложенными вариантами
        specifications = {}
        try:
            if main_product.specifications:
                specifications = json.loads(main_product.specifications)
        except json.JSONDecodeError:
            pass
        
        variants = []
        
        # Извлекаем варианты из specifications.variants
        if 'variants' in specifications:
            # Сортируем варианты по цвету (в алфавитном порядке)
            sorted_variants = sorted(specifications['variants'], key=lambda x: x.get('specifications', {}).get('color', ''))
            
            for variant_info in sorted_variants:
                # Находим цену для этого варианта по SKU
                variant_product = db.query(Product).filter(Product.sku == variant_info['sku']).first()
                price = None
                if variant_product:
                    price = db.query(CurrentPrice).filter(CurrentPrice.sku == variant_product.sku).first()
                
                variant_specs = variant_info.get('specifications', {})
                
                # Получаем изображения для цвета из основного продукта
                variant_color = variant_specs.get('color', '')
                variant_color_normalized = variant_color.lower().replace(' ', '-')
                variant_images = []
                
                if 'images' in specifications:
                    for img_info in specifications['images']:
                        if isinstance(img_info, dict) and 'color' in img_info:
                            img_color = img_info.get('color', '').lower()
                            if img_color == variant_color_normalized:
                                variant_images.append(img_info.get('url', ''))
                        elif isinstance(img_info, str):
                            # Прямой URL - пытаемся определить цвет из пути
                            img_path = img_info.lower()
                            if variant_color_normalized in img_path:
                                variant_images.append(img_info)
                
                variant_data = {
                    "sku": variant_info['sku'],
                    "name": variant_info['name'],
                    "price": price.price if price else 0.0,
                    "old_price": price.old_price if price else 0.0,
                    "discount_percentage": price.discount_percentage if price else 0.0,
                    "currency": price.currency if price else "RUB",
                    "stock": variant_info.get('stock', 0),
                    "is_available": variant_info.get('is_available', True),
                    
                    # Добавляем спецификации варианта
                    "color": variant_specs.get('color', ''),
                    "memory": variant_specs.get('memory', ''),
                    "sim_type": variant_specs.get('sim_type', ''),
                    
                    # Изображения для этого цвета
                    "images": variant_images,
                    "main_image": variant_images[0] if variant_images else ""
                }
                
                variants.append(variant_data)
        
        return {
            "model": model,
            "variants": variants,
            "total_variants": len(variants)
        }
    
    # Fallback: Если основного продукта нет, найдем все варианты для данной модели
    variants_query = db.query(Product, CurrentPrice).outerjoin(
        CurrentPrice, Product.sku == CurrentPrice.sku
    ).filter(Product.level_2 == model).order_by(Product.specifications)  # Сортируем по specifications
    
    variants = []
    for result in variants_query.all():
        product = result[0]
        price = result[1]
        
        # Получаем спецификации варианта
        specifications = {}
        try:
            if product.specifications:
                specifications = json.loads(product.specifications)
        except json.JSONDecodeError:
            pass
        
        # Получаем изображения
        images = get_product_images(product, db)
        
        # Используем поля из модели Product напрямую, fallback на specifications
        color = product.color if product.color else specifications.get('color', '')
        memory = product.disk if product.disk else specifications.get('disk', specifications.get('memory', ''))
        sim_type = product.sim_config if product.sim_config else specifications.get('sim_config', specifications.get('sim_type', ''))
        
        variant_data = {
            "sku": product.sku,
            "name": product.name,
            "price": price.price if price else 0.0,
            "old_price": price.old_price if price else 0.0,
            "discount_percentage": price.discount_percentage if price else 0.0,
            "currency": price.currency if price else "RUB",
            "stock": product.stock,
            "is_available": product.is_available,
            
            # Добавляем спецификации варианта (цвет, память, SIM)
            "color": color,
            "memory": memory,
            "sim_type": sim_type,
            
            # Изображения для этого варианта (один цвет обычно)
            "images": images,
            "main_image": images[0] if images else ""
        }
        
        variants.append(variant_data)
    
    # Сортируем варианты по цвету для fallback случая
    variants.sort(key=lambda x: x.get('color', ''))
    
    return {
        "model": model,
        "variants": variants,
        "total_variants": len(variants)
    }

@app.get("/products/{product_id}", response_model=ProductDetailResponse)
async def get_product(product_id: int, db: Session = Depends(get_db)):
    """Get detailed product information"""
    result = db.query(Product, CurrentPrice).join(
        CurrentPrice, Product.sku == CurrentPrice.sku
    ).filter(
        Product.id == product_id
    ).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product, price = result
    
    try:
        specifications = json.loads(product.specifications) if product.specifications else {}
    except json.JSONDecodeError:
        specifications = {}
    
    # Получаем описание из level2_descriptions
    desc = ""
    level2_specs = {}
    if product.level_2:
        level2_desc = db.query(Level2Description).filter(Level2Description.level_2 == product.level_2).first()
        if level2_desc:
            desc = level2_desc.description or ""
            if level2_desc.details:
                try:
                    level2_specs = json.loads(level2_desc.details) if isinstance(level2_desc.details, str) else level2_desc.details
                except json.JSONDecodeError:
                    level2_specs = {}
    
    # Объединяем характеристики из level2_descriptions с существующими specifications
    all_specifications = {**level2_specs, **specifications}
    
    images = get_product_images(product, db)
    
    return ProductDetailResponse(
        id=product.id,
        sku=product.sku,
        name=product.name,
        description=desc,
        brand=product.brand,
        model=product.level_2 or "",
        category_name=product.level_0 or "Без категории",
        image_url=images[0] if images else '',
        images=images,
        specifications=all_specifications,
        price=price.price,
        old_price=price.old_price,
        discount_percentage=price.discount_percentage,
        currency=price.currency,
        is_available=product.is_available,
        created_at=product.created_at.isoformat()
    )

@app.get("/search")
async def search_products(
    q: str,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Search products by name, brand, or level_2 - returns unique models only"""
    search_term = f"%{q}%"
    
    # Создаем фильтры для поиска
    search_filters = [
        Product.is_available == True,
        (
            Product.name.ilike(search_term) |
            Product.brand.ilike(search_term) |
            Product.level_2.ilike(search_term)
        )
    ]
    
    # Используем точно такую же логику как в get_products - получаем уникальные модели
    subquery = db.query(
        func.min(Product.id).label('id')
    ).filter(*search_filters).group_by(Product.level_2, Product.brand).subquery()
    
    # Теперь получаем только те товары, которые являются представителями групп
    final_query = db.query(Product, CurrentPrice).outerjoin(
        CurrentPrice, Product.sku == CurrentPrice.sku
    ).filter(
        Product.id == subquery.c.id
    ).order_by(Product.level_2, Product.id)
    
    # Применяем лимит
    results = final_query.limit(limit).all()
    
    products = []
    for product, price in results:
        try:
            specifications = json.loads(product.specifications) if product.specifications else {}
        except json.JSONDecodeError:
            specifications = {}
        
        # Получаем описание из level2_descriptions
        desc = ""
        level2_specs = {}
        if product.level_2:
            level2_desc = db.query(Level2Description).filter(Level2Description.level_2 == product.level_2).first()
            if level2_desc:
                desc = level2_desc.description or ""
                if level2_desc.details:
                    try:
                        level2_specs = json.loads(level2_desc.details) if isinstance(level2_desc.details, str) else level2_desc.details
                    except json.JSONDecodeError:
                        level2_specs = {}
        
        # Объединяем характеристики из level2_descriptions с существующими specifications
        all_specifications = {**level2_specs, **specifications}
        
        images = get_product_images(product, db)
        
        # Получаем название категории из level полей
        category_name = product.level_0 or "Без категории"
        if product.level_1:
            category_name += f" / {product.level_1}"
        if product.level_2:
            category_name += f" / {product.level_2}"
        
        # Если цена отсутствует, используем значения по умолчанию (аналогично get_products)
        if price is None:
            price_obj = type('Price', (), {
                'price': 0.0,
                'old_price': 0.0,
                'discount_percentage': 0.0,
                'currency': 'RUB'
            })()
        else:
            price_obj = price
        
        products.append(ProductResponse(
            id=product.id,
            sku=product.sku,
            name=product.name,
            description=desc,
            brand=product.brand,
            model=product.level_2 or "",
            category_name=category_name,
            level_2=product.level_2,
            image_url=images[0] if images else '',
            images=images,
            specifications=specifications,
            price=price_obj.price,
            old_price=price_obj.old_price,
            discount_percentage=price_obj.discount_percentage,
            currency=price_obj.currency,
        ))
    
    return products

@app.get("/webapp")
async def webapp():
    """Serve the web app"""
    return FileResponse("webapp.html")

@app.get("/login")
async def login_page():
    return FileResponse("login.html")

@app.post("/login")
async def login(request: Request, response: Response):
    try:
        body = await request.json()
    except Exception:
        body = {}
    username = (body.get("username") or "").strip()
    password = (body.get("password") or "").strip()

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        # Set secure httpOnly cookie
        response = JSONResponse({"success": True, "message": "Authenticated"})
        response.set_cookie(
            key="admin_session",
            value=ADMIN_SESSION_TOKEN,
            httponly=True,
            max_age=60 * 60 * 8,  # 8 hours
            samesite="lax"
        )
        return response
    raise HTTPException(status_code=401, detail="Invalid credentials")

@app.post("/logout")
async def logout(response: Response):
    res = JSONResponse({"success": True})
    res.delete_cookie("admin_session")
    return res

@app.get("/admin")
async def admin(request: Request):
    """Serve the admin panel (protected)"""
    if not is_admin_authenticated(request):
        return RedirectResponse(url="/login", status_code=302)
    return FileResponse("admin.html")

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}

# Excel Management API
@app.get("/api/excel/template/products")
async def download_products_template():
    """Скачать шаблон Excel файла для добавления товаров"""
    excel_handler = ExcelHandler()
    template_data = excel_handler.create_products_template()
    
    return StreamingResponse(
        io.BytesIO(template_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_template.xlsx"}
    )

@app.get("/api/excel/template/prices")
async def download_prices_template():
    """Скачать шаблон Excel файла для обновления цен"""
    excel_handler = ExcelHandler()
    template_data = excel_handler.create_prices_template()
    
    return StreamingResponse(
        io.BytesIO(template_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prices_template.xlsx"}
    )

@app.post("/api/excel/import/products")
async def import_products_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Импортировать товары из Excel файла"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel (.xlsx или .xls)")
    
    try:
        # Читаем содержимое файла
        file_content = await file.read()
        
        # Парсим Excel файл
        excel_handler = ExcelHandler()
        products_data = excel_handler.parse_products_excel(file_content)
        
        # Также парсим изображения если есть лист "Изображения"
        images_data = []
        try:
            images_data = excel_handler.parse_images_excel(file_content)
        except Exception as e:
            print(f"Предупреждение: Не удалось загрузить изображения: {e}")
        
        # Добавляем товары в базу данных
        added_count = 0
        errors = []
        
        for i, product_data in enumerate(products_data):
            try:
                # Проверяем, что указаны обязательные поля level0
                if not product_data.get('level0'):
                    errors.append(f"Товар {i+1}: Не указана основная категория (level0)")
                    continue
                
                # Проверяем обязательные поля для генерации SKU
                if not product_data.get('brand') or not product_data.get('level2'):
                    errors.append(f"Товар {i+1}: Не указан бренд или модель, необходимые для генерации SKU")
                    continue
                
                # Генерируем временный уникальный SKU если не указан
                if product_data.get('sku'):
                    # Проверяем уникальность SKU
                    existing = db.query(Product).filter(Product.sku == product_data['sku']).first()
                    if existing:
                        errors.append(f"Товар {i+1}: SKU '{product_data['sku']}' уже существует")
                        continue
                    sku = product_data['sku']
                else:
                    # Генерируем уникальный SKU на основе данных и timestamp
                    import time
                    timestamp = int(time.time() * 1000) % 100000  # последние 5 цифр timestamp
                    sku = f"{product_data['brand'][:3].upper()}{product_data['level2'][:5].upper()}{timestamp}"
                
                # Создать товар
                parsed_images = parse_images_from_string(product_data['image_url'])
                
                # Сформировать specifications из данных
                specs = dict(product_data.get('specifications') or {})
                if product_data.get('color'):
                    specs['color'] = product_data['color']
                if product_data.get('ram'):
                    specs['ram'] = product_data['ram']
                if product_data.get('disk'):
                    specs['disk'] = product_data['disk']
                if product_data.get('sim_config'):
                    specs['sim_config'] = product_data['sim_config']

                db_product = Product(
                    sku=sku,
                    name=product_data['name'],
                    level_0=product_data['level0'],
                    level_1=product_data.get('level1'),
                    level_2=product_data.get('level2'),
                    brand=product_data['brand'],
                    specifications=json.dumps(specs),
                    stock=product_data['stock'],
                    is_available=True
                )
                
                db.add(db_product)
                db.flush()  # Получить ID

                # Ensure categories exist
                ensure_category_exists(db, product_data.get('level0'), product_data.get('level1'), product_data.get('level2'))
                
                # Создать начальную цену
                db_price = CurrentPrice(
                    sku=sku,
                    price=product_data['price'],
                    old_price=product_data['price'],
                    discount_percentage=0.0,
                    currency=product_data['currency'],
                    updated_at=datetime.utcnow()
                )
                
                db.add(db_price)
                
                # Создать запись изображений в ProductImage если есть изображения
                if parsed_images and product_data.get('level2') and product_data.get('color'):
                    product_image = ProductImage(
                        level_2=product_data['level2'],
                        color=product_data['color'],
                        img_list=json.dumps(parsed_images)
                    )
                    db.add(product_image)
                
                added_count += 1
                
            except Exception as e:
                errors.append(f"Товар {i+1}: {str(e)}")
        
        db.commit()
        
        # Обрабатываем изображения если они есть
        images_added = 0
        images_updated = 0
        images_errors = []
        
        for image_data in images_data:
            try:
                # Ищем существующую запись
                existing_image = db.query(ProductImage).filter(
                    ProductImage.level_2 == image_data['level_2'],
                    ProductImage.color == image_data['color']
                ).first()
                
                # Конвертируем список изображений в JSON
                img_list_json = json.dumps(image_data['img_list'])
                
                if existing_image:
                    # Обновляем существующую запись
                    existing_image.img_list = img_list_json
                    images_updated += 1
                else:
                    # Создаем новую запись
                    new_image = ProductImage(
                        level_2=image_data['level_2'],
                        color=image_data['color'],
                        img_list=img_list_json
                    )
                    db.add(new_image)
                    images_added += 1
                    
            except Exception as e:
                images_errors.append(f"Ошибка при обработке изображений {image_data['level_2']} - {image_data['color']}: {str(e)}")
        
        # Сохраняем изменения изображений
        if images_data:
            db.commit()
        
        return {
            "message": "Импорт завершен",
            "added": added_count,
            "errors": errors,
            "total_processed": len(products_data),
            "images_added": images_added,
            "images_updated": images_updated,
            "images_errors": images_errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при импорте: {str(e)}")

@app.post("/api/excel/import/prices")
async def import_prices_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Обновить цены из Excel файла"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel (.xlsx или .xls)")
    
    try:
        # Читаем содержимое файла
        file_content = await file.read()
        
        # Парсим Excel файл
        excel_handler = ExcelHandler()
        prices_data = excel_handler.parse_prices_excel(file_content)
        
        # Обновляем цены в базе данных через ручной менеджер
        updated_count = 0
        errors = []
        
        for i, price_data in enumerate(prices_data):
            try:
                success = manual_price_manager.update_price_from_excel_data(price_data, db)
                if success:
                    updated_count += 1
                else:
                    errors.append(f"Строка {i+1}: Не удалось обновить цену")
            except Exception as e:
                errors.append(f"Строка {i+1}: {str(e)}")
        
        db.commit()
        
        return {
            "message": "Обновление цен завершено",
            "updated": updated_count,
            "errors": errors,
            "total_processed": len(prices_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при обновлении цен: {str(e)}")

@app.post("/api/excel/update-or-create/products")
async def update_or_create_products_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Массовое обновление существующих товаров (по SKU) или добавление новых"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel (.xlsx или .xls)")
    
    try:
        # Читаем содержимое файла
        file_content = await file.read()
        
        # Парсим Excel файл
        excel_handler = ExcelHandler()
        products_data = excel_handler.parse_products_excel(file_content)
        
        # Также парсим изображения если есть лист "Изображения"
        images_data = []
        try:
            images_data = excel_handler.parse_images_excel(file_content)
        except Exception as e:
            print(f"Предупреждение: Не удалось загрузить изображения: {e}")
        
        added_count = 0
        updated_count = 0
        errors = []
        
        for i, product_data in enumerate(products_data):
            try:
                # Генерируем SKU если не указан
                if not product_data.get('sku'):
                    import time
                    timestamp = int(time.time() * 1000) % 100000
                    brand_part = product_data.get('brand', 'UNK')[:3].upper()
                    model_part = product_data.get('level_2', 'UNK')[:5].upper()
                    product_data['sku'] = f"{brand_part}{model_part}{timestamp}"
                
                # Проверяем, существует ли товар с таким SKU
                existing_product = db.query(Product).filter(Product.sku == product_data['sku']).first()
                
                if existing_product:
                    # Обновляем существующий товар
                    existing_product.name = product_data['name']
                    # existing_product.description = product_data.get('description', '')  # поле удалено
                    existing_product.level_0 = product_data['level0']
                    existing_product.level_1 = product_data.get('level1', '')
                    existing_product.level_2 = product_data.get('level2', '')
                    existing_product.brand = product_data.get('brand', '')
                    existing_product.stock = product_data.get('stock', 0)
                    # Ensure categories exist for updated levels
                    ensure_category_exists(db, product_data.get('level0'), product_data.get('level1'), product_data.get('level2'))
                    
                    # Обновляем характеристики в specifications JSON
                    try:
                        existing_specs = json.loads(existing_product.specifications) if existing_product.specifications else {}
                    except json.JSONDecodeError:
                        existing_specs = {}
                    for key in ['color', 'disk', 'ram', 'sim_config']:
                        if product_data.get(key):
                            existing_specs[key] = product_data[key]
                    existing_product.specifications = json.dumps(existing_specs)
                    
                    # Обновляем изображения если указаны (в таблице ProductImage)
                    if product_data.get('image_url') and existing_product.level_2 and existing_product.color:
                        parsed_images = parse_images_from_string(product_data['image_url'])
                        img_list_json = json.dumps(parsed_images) if parsed_images else '[]'
                        
                        # Ищем или создаем запись в ProductImage
                        product_image = db.query(ProductImage).filter(
                            ProductImage.level_2 == existing_product.level_2,
                            ProductImage.color == existing_product.color
                        ).first()
                        
                        if product_image:
                            product_image.img_list = img_list_json
                        else:
                            product_image = ProductImage(
                                level_2=existing_product.level_2,
                                color=existing_product.color,
                                img_list=img_list_json
                            )
                            db.add(product_image)
                    
                    # Обновляем цену
                    current_price = db.query(CurrentPrice).filter(CurrentPrice.sku == product_data['sku']).first()
                    if current_price:
                        current_price.price = product_data['price']
                        current_price.currency = product_data.get('currency', 'RUB')
                        current_price.last_updated = datetime.now()
                    else:
                        # Создаем новую цену
                        new_price = CurrentPrice(
                            sku=product_data['sku'],
                            price=product_data['price'],
                            old_price=product_data['price'],
                            currency=product_data.get('currency', 'RUB')
                        )
                        db.add(new_price)
                    
                    updated_count += 1
                else:
                    # Создаем новый товар
                    parsed_images = parse_images_from_string(product_data.get('image_url', ''))
                    
                    specs = dict(product_data.get('specifications') or {})
                    if product_data.get('color'):
                        specs['color'] = product_data['color']
                    if product_data.get('ram'):
                        specs['ram'] = product_data['ram']
                    if product_data.get('disk'):
                        specs['disk'] = product_data['disk']
                    if product_data.get('sim_config'):
                        specs['sim_config'] = product_data['sim_config']

                    db_product = Product(
                        sku=product_data['sku'],
                        name=product_data['name'],
                        level_0=product_data['level0'],
                        level_1=product_data.get('level1', ''),
                        level_2=product_data.get('level2', ''),
                        brand=product_data.get('brand', ''),
                        stock=product_data.get('stock', 0),
                        specifications=json.dumps(specs),
                        is_available=True
                    )
                    db.add(db_product)
                    # Ensure categories exist
                    ensure_category_exists(db, product_data.get('level0'), product_data.get('level1'), product_data.get('level2'))
                    
                    # Добавляем цену
                    db_price = CurrentPrice(
                        sku=product_data['sku'],
                        price=product_data['price'],
                        old_price=product_data['price'],
                        currency=product_data.get('currency', 'RUB')
                    )
                    db.add(db_price)
                    
                    # Создать запись изображений в ProductImage если есть изображения
                    if parsed_images and product_data.get('level_2') and product_data.get('color'):
                        product_image = ProductImage(
                            level_2=product_data['level_2'],
                            color=product_data['color'],
                            img_list=json.dumps(parsed_images)
                        )
                        db.add(product_image)
                    
                    added_count += 1
                    
            except Exception as e:
                errors.append(f"Строка {i+2}: {str(e)}")
        
        db.commit()
        
        # Обрабатываем изображения если они есть
        images_added = 0
        images_updated = 0
        images_errors = []
        
        for image_data in images_data:
            try:
                # Ищем существующую запись
                existing_image = db.query(ProductImage).filter(
                    ProductImage.level_2 == image_data['level_2'],
                    ProductImage.color == image_data['color']
                ).first()
                
                # Конвертируем список изображений в JSON
                img_list_json = json.dumps(image_data['img_list'])
                
                if existing_image:
                    # Обновляем существующую запись
                    existing_image.img_list = img_list_json
                    images_updated += 1
                else:
                    # Создаем новую запись
                    new_image = ProductImage(
                        level_2=image_data['level_2'],
                        color=image_data['color'],
                        img_list=img_list_json
                    )
                    db.add(new_image)
                    images_added += 1
                    
            except Exception as e:
                images_errors.append(f"Ошибка при обработке изображений {image_data['level_2']} - {image_data['color']}: {str(e)}")
        
        # Сохраняем изменения изображений
        if images_data:
            db.commit()
        
        return {
            "success": True,
            "message": f"Обработка завершена: добавлено {added_count}, обновлено {updated_count}",
            "added": added_count,
            "updated": updated_count,
            "errors": errors,
            "total_processed": len(products_data),
            "images_added": images_added,
            "images_updated": images_updated,
            "images_errors": images_errors
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка при обработке файла: {str(e)}")

@app.post("/api/excel/import/images")
async def import_images_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Импортировать изображения из Excel файла"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel (.xlsx или .xls)")
    
    try:
        # Читаем содержимое файла
        file_content = await file.read()
        
        # Парсим Excel файл
        excel_handler = ExcelHandler()
        images_data = excel_handler.parse_images_excel(file_content)
        
        # Добавляем изображения в базу данных
        added_count = 0
        updated_count = 0
        errors = []
        
        for image_data in images_data:
            try:
                # Ищем существующую запись
                existing_image = db.query(ProductImage).filter(
                    ProductImage.level_2 == image_data['level_2'],
                    ProductImage.color == image_data['color']
                ).first()
                
                # Конвертируем список изображений в JSON
                img_list_json = json.dumps(image_data['img_list'])
                
                if existing_image:
                    # Обновляем существующую запись
                    existing_image.img_list = img_list_json
                    updated_count += 1
                else:
                    # Создаем новую запись
                    new_image = ProductImage(
                        level_2=image_data['level_2'],
                        color=image_data['color'],
                        img_list=img_list_json
                    )
                    db.add(new_image)
                    added_count += 1
                    
            except Exception as e:
                errors.append(f"Ошибка при обработке {image_data['level_2']} - {image_data['color']}: {str(e)}")
        
        # Сохраняем изменения
        db.commit()
        
        return {
            "success": True,
            "message": f"Обработка изображений завершена: добавлено {added_count}, обновлено {updated_count}",
            "added": added_count,
            "updated": updated_count,
            "errors": errors,
            "total_processed": len(images_data)
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка при обработке файла: {str(e)}")

@app.get("/download-price-template")
async def download_price_template(db: Session = Depends(get_db)):
    """Скачать простой шаблон Excel для обновления цен: SKU - новая цена - старая цена"""
    try:
        from io import BytesIO
        
        # Получаем существующие товары с их SKU и текущими ценами
        products = db.query(Product, CurrentPrice).outerjoin(
            CurrentPrice, Product.sku == CurrentPrice.sku
        ).filter(Product.is_available == True).limit(50).all()
        
        # Создаем DataFrame с примерами
        data = []
        if products:
            for product, price in products:
                data.append({
                    'SKU': product.sku,
                    'Новая цена': price.price if price else 0.0,
                    'Старая цена': price.old_price if price and price.old_price else (price.price if price else 0.0)
                })
        else:
            # Примеры если нет товаров
            data = [
                {'SKU': 'APPIP16', 'Новая цена': 59990, 'Старая цена': 69990},
                {'SKU': 'APPIP16PRO', 'Новая цена': 89990, 'Старая цена': 99990},
                {'SKU': 'SAM-GALAXY-S24', 'Новая цена': 79990, 'Старая цена': 89990}
            ]
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Цены')
            
            # Стилизация
            workbook = writer.book
            worksheet = writer.sheets['Цены']
            
            # Стили для заголовков
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        output.seek(0)
        
        return StreamingResponse(
            BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=prices_template.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка создания шаблона: {str(e)}")

@app.post("/import-prices")
async def import_prices_simple(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Простое обновление цен из Excel: SKU - новая цена - старая цена"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel (.xlsx или .xls)")
    
    try:
        # Читаем содержимое файла
        file_content = await file.read()
        
        # Парсим как DataFrame
        df = pd.read_excel(io.BytesIO(file_content))
        
        # Проверяем наличие нужных колонок
        if len(df.columns) < 3:
            raise HTTPException(status_code=400, detail="Файл должен содержать минимум 3 колонки: SKU, Новая цена, Старая цена")
        
        # Используем первые 3 колонки
        sku_col = df.columns[0]
        new_price_col = df.columns[1] 
        old_price_col = df.columns[2]
        
        updated_count = 0
        errors = []
        not_found = []
        
        for index, row in df.iterrows():
            try:
                # Пропускаем пустые строки
                if pd.isna(row[sku_col]) or pd.isna(row[new_price_col]):
                    continue
                
                sku = str(row[sku_col]).strip()
                try:
                    new_price = float(row[new_price_col])
                except:
                    errors.append(f"Строка {index + 2}: Неверный формат новой цены")
                    continue
                
                try:
                    old_price = float(row[old_price_col]) if not pd.isna(row[old_price_col]) else new_price
                except:
                    old_price = new_price
                
                # Найти товар по SKU
                product = db.query(Product).filter(Product.sku == sku).first()
                
                if not product:
                    not_found.append(f"Строка {index + 2}: Товар с SKU '{sku}' не найден")
                    continue
                
                # Обновить или создать цену
                current_price = db.query(CurrentPrice).filter(CurrentPrice.sku == product.sku).first()
                
                if current_price:
                    # Обновить существующую цену
                    current_price.price = new_price
                    current_price.old_price = old_price
                    current_price.discount_percentage = ((old_price - new_price) / old_price * 100) if old_price > new_price else 0
                else:
                    # Создать новую цену
                    new_price_obj = CurrentPrice(
                        product_id=product.id,
                        price=new_price,
                        old_price=old_price,
                        discount_percentage=((old_price - new_price) / old_price * 100) if old_price > new_price else 0,
                        currency='RUB'
                    )
                    db.add(new_price_obj)
                
                updated_count += 1
                
            except Exception as e:
                errors.append(f"Строка {index + 2}: {str(e)}")
        
        db.commit()
        
        return {
            "message": "Обновление цен завершено",
            "updated": updated_count,
            "errors": errors,
            "not_found": not_found,
            "total_processed": len(df)
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при обновлении цен: {str(e)}")

@app.get("/api/excel/export/products")
async def export_products_to_excel(db: Session = Depends(get_db)):
    """Экспортировать все товары в формате для редактирования и повторного импорта"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Получить все товары с ценами
        results = db.query(Product, CurrentPrice).outerjoin(
            CurrentPrice, Product.sku == CurrentPrice.sku
        ).all()
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        # Заголовки (те же что в шаблоне для импорта, но без изображений)
        headers = [
            'SKU товара', 'Название товара*', 'Описание', 
            'Основная категория (level0)*', 'Подкатегория (level1)*', 'Детальная категория (level2)*',
            'Бренд', 'Цена*', 'Валюта', 'Количество на складе',
            'Характеристики (JSON)'
        ]
        
        # Добавляем заголовки со стилем
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавляем данные товаров
        for row_idx, (product, price) in enumerate(results, 2):
            # Получаем характеристики
            try:
                specifications = json.loads(product.specifications) if product.specifications else {}
                specs_str = json.dumps(specifications, ensure_ascii=False) if specifications else ''
            except:
                specs_str = ''
            
            # Заполняем строку данными (без столбца изображений)
            ws.cell(row=row_idx, column=1, value=product.sku or '')
            ws.cell(row=row_idx, column=2, value=product.name or '')
            ws.cell(row=row_idx, column=3, value='')  # description поле удалено
            ws.cell(row=row_idx, column=4, value=product.level_0 or '')
            ws.cell(row=row_idx, column=5, value=product.level_1 or '')
            ws.cell(row=row_idx, column=6, value=product.level_2 or '')
            ws.cell(row=row_idx, column=7, value=product.brand or '')
            ws.cell(row=row_idx, column=8, value=price.price if price else 0.0)
            ws.cell(row=row_idx, column=9, value=price.currency if price else 'RUB')
            ws.cell(row=row_idx, column=10, value=product.stock or 0)
            ws.cell(row=row_idx, column=11, value=specs_str)
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем лист с изображениями
        images_ws = wb.create_sheet("Изображения")
        images_headers = [
            'Модель (level_2)*', 'Цвет*', 'URL изображений (через запятую)*'
        ]
        
        # Добавляем заголовки для изображений
        for col, header in enumerate(images_headers, 1):
            cell = images_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Получаем все изображения из таблицы ProductImage
        product_images = db.query(ProductImage).all()
        
        # Добавляем данные изображений
        for row_idx, product_image in enumerate(product_images, 2):
            try:
                # Парсим JSON с изображениями
                images_data = json.loads(product_image.img_list) if product_image.img_list else []
                
                # Проверяем, что images_data это список строк, а не список словарей
                if images_data and isinstance(images_data[0], dict):
                    # Если это список словарей, извлекаем URL
                    image_urls = ', '.join([img.get('url', '') for img in images_data if img.get('url')])
                else:
                    # Если это список строк
                    image_urls = ', '.join(images_data) if images_data else ''
                
                images_ws.cell(row=row_idx, column=1, value=product_image.level_2 or '')
                images_ws.cell(row=row_idx, column=2, value=product_image.color or '')
                images_ws.cell(row=row_idx, column=3, value=image_urls)
            except Exception as e:
                print(f"Ошибка при обработке изображений для {product_image.level_2} - {product_image.color}: {e}")
                continue
        
        # Автоподбор ширины колонок для изображений
        for column in images_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            images_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=current_products.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при экспорте: {str(e)}")

# Additional Price Management API
@app.get("/api/prices/current")
async def get_current_prices():
    """Получить все текущие цены"""
    try:
        prices = manual_price_manager.get_all_current_prices()
        return {
            "message": "Текущие цены получены",
            "prices": prices,
            "total": len(prices)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения цен: {str(e)}")

@app.get("/api/prices/history/{product_id}")
async def get_price_history(product_id: int, limit: int = 10):
    """Получить историю цен товара"""
    try:
        history = manual_price_manager.get_price_history(product_id, limit)
        return {
            "message": f"История цен товара {product_id}",
            "product_id": product_id,
            "history": history,
            "total": len(history)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения истории цен: {str(e)}")

@app.post("/api/prices/update-single")
async def update_single_price(
    product_id: int,
    new_price: float,
    currency: str = "RUB",
    db: Session = Depends(get_db)
):
    """Обновить цену одного товара"""
    try:
        price_data = {
            'product_id': product_id,
            'price': new_price,
            'currency': currency
        }
        
        success = manual_price_manager.update_price_from_excel_data(price_data, db)
        
        if success:
            db.commit()
            return {
                "message": "Цена успешно обновлена",
                "product_id": product_id,
                "new_price": new_price,
                "currency": currency
            }
        else:
            raise HTTPException(status_code=400, detail="Не удалось обновить цену")
            
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка обновления цены: {str(e)}")

@app.get("/product-images/{model_key}/{color}")
async def get_product_images_by_color(model_key: str, color: str, db: Session = Depends(get_db)):
    """Get images for a specific product color from ProductImage table"""
    import os
    import urllib.parse
    
    # Декодируем URL параметры
    model_key = urllib.parse.unquote(model_key)
    color = urllib.parse.unquote(color)
    
    # Нормализуем названия для поиска в БД
    # model_key может быть как "iphone17pro", так и "iPhone 17 Pro"
    # Попробуем найти в БД по разным вариантам
    
    # Вариант 1: Ищем напрямую как level_2
    product_image = db.query(ProductImage).filter(
        ProductImage.level_2 == model_key,
        ProductImage.color == color
    ).first()
    
    # Вариант 2: Если не нашли, пробуем нормализованные варианты
    if not product_image:
        # Попробуем найти похожий level_2 и color (case-insensitive)
        all_images = db.query(ProductImage).all()
        for img in all_images:
            # Нормализуем level_2: убираем пробелы и дефисы, приводим к нижнему регистру
            img_level2_normalized = img.level_2.lower().replace(' ', '').replace('-', '')
            model_key_normalized = model_key.lower().replace(' ', '').replace('-', '')
            
            # Нормализуем color: убираем пробелы и дефисы, приводим к нижнему регистру
            img_color_normalized = img.color.lower().replace(' ', '').replace('-', '')
            color_normalized = color.lower().replace(' ', '').replace('-', '')
            
            if (img_level2_normalized == model_key_normalized and
                img_color_normalized == color_normalized):
                product_image = img
                break
    
    # Если нашли в БД - возвращаем из базы
    if product_image and product_image.img_list:
        try:
            images_data = json.loads(product_image.img_list)
            image_paths = []
            
            # Обрабатываем массив изображений
            for img_data in images_data:
                if isinstance(img_data, dict):
                    # Формат: {"url": "...", "alt": "..."}
                    image_paths.append(img_data["url"])
                elif isinstance(img_data, str):
                    # Формат: просто строка с URL
                    image_paths.append(img_data)
            
            return {"image_paths": image_paths}
        except (json.JSONDecodeError, KeyError) as e:
            # Если ошибка парсинга, пробуем fallback на файловую систему
            pass
    
    # Fallback: ищем в файловой системе (старая логика)
    try:
        actual_model_key = normalize_model_key(model_key)
        normalized_color = normalize_color_name(color)
        
        image_folder = f"static/images/products/{actual_model_key}/{normalized_color}"
        
        if not os.path.exists(image_folder):
            raise HTTPException(
                status_code=404, 
                detail=f"Изображения не найдены: model='{model_key}', color='{color}' (поиск в БД и файловой системе)"
            )
        
        # Формируем пути к реальным файлам
        image_paths = []
        all_files = os.listdir(image_folder)
        jpg_files = [f for f in all_files if f.endswith('.jpg')]
        jpg_files.sort()
        
        for file_name in jpg_files:
            image_path = f"/static/images/products/{actual_model_key}/{normalized_color}/{file_name}"
            image_paths.append(image_path)
                
        return {"image_paths": image_paths}
        
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Изображения не найдены: {str(e)}")

@app.get("/color-schemes/{model_key}")
async def get_color_schemes(model_key: str, db: Session = Depends(get_db)):
    """Get color schemes for a product from database"""
    from models import ModelColorScheme
    
    # Ищем схему цветов в базе данных
    color_scheme = db.query(ModelColorScheme).filter(
        ModelColorScheme.model_key == model_key
    ).first()
    
    if not color_scheme:
        raise HTTPException(status_code=404, detail=f"Цветовая схема не найдена для {model_key}")
    
    try:
        # Парсим JSON с цветами
        colors_data = json.loads(color_scheme.colors_json)
        
        # Определяем цвет по умолчанию (первый в списке)
        default_color = colors_data[0]["value"] if colors_data else ""
        
        return {
            "colors": colors_data,
            "default_color": default_color
        }
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Ошибка парсинга цветовой схемы")

@app.get("/variant-schemes/{model_key}")
async def get_variant_schemes(model_key: str, db: Session = Depends(get_db)):
    """Get variant schemes for a product from database"""
    from models import ModelVariantScheme
    
    # Ищем схему вариантов в базе данных
    variant_scheme = db.query(ModelVariantScheme).filter(
        ModelVariantScheme.model_key == model_key
    ).first()
    
    if not variant_scheme:
        raise HTTPException(status_code=404, detail=f"Схема вариантов не найдена для {model_key}")
    
    try:
        # Парсим JSON с вариантами
        variants_data = json.loads(variant_scheme.variants_json)
        
        return {
            "variants": variants_data
        }
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Ошибка парсинга схемы вариантов")

# Новые endpoints для иерархической фильтрации

@app.get("/hierarchy/brands")
async def get_brands(db: Session = Depends(get_db)):
    """Получить все доступные бренды"""
    brands = db.query(Product.brand.distinct()).filter(
        Product.brand.isnot(None),
        Product.is_available == True
    ).all()
    return [brand[0] for brand in brands]

@app.get("/hierarchy/levels")
async def get_hierarchy_levels(
    level: Optional[int] = None,
    brand: Optional[str] = None,
    parent_level0: Optional[str] = None,
    parent_level1: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Получить значения уровней иерархии
    
    level: кавой уровень получить (0, 1, или 2)
    brand: фильтр по бренду
    parent_level0: для уровня 1 - фильтр по level0
    parent_level1: для уровня 2 - фильтр по level1
    """
    
    filters = [Product.is_available == True]
    
    if brand:
        filters.append(Product.brand == brand)
    if parent_level0:
        filters.append(Product.level_0 == parent_level0)
    if parent_level1:
        filters.append(Product.level_1 == parent_level1)
    
    query = db.query(Product).filter(and_(*filters))
    
    if level == 0:
        # Получаем все level_0
        results = query.with_entities(Product.level_0.distinct()).filter(
            Product.level_0.isnot(None)
        ).all()
        return [item[0] for item in results if item[0]]
        
    elif level == 1:
        # Получаем все level_1
        results = query.with_entities(Product.level_1.distinct()).filter(
            Product.level_1.isnot(None)
        ).all()
        return [item[0] for item in results if item[0]]
        
    elif level == 2:
        # Получаем все level_2
        results = query.with_entities(Product.level_2.distinct()).filter(
            Product.level_2.isnot(None)
        ).all()
        return [item[0] for item in results if item[0]]
    
    else:
        # Возвращаем всю иерархию
        return {
            "level0": db.query(Product.level_0.distinct()).filter(
                Product.level_0.isnot(None),
                Product.is_available == True
            ).all(),
            "level1": db.query(Product.level_1.distinct()).filter(
                Product.level_1.isnot(None),
                Product.is_available == True
            ).all(),
            "level2": db.query(Product.level_2.distinct()).filter(
                Product.level_2.isnot(None),
                Product.is_available == True
            ).all()
        }

@app.get("/hierarchy/models")
async def get_models(
    brand: Optional[str] = None,
    level0: Optional[str] = None,
    level1: Optional[str] = None,
    level2: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Получить все доступные модели с фильтрацией"""
    
    filters = [Product.is_available == True]
    
    if brand:
        filters.append(Product.brand == brand)
    if level0:
        filters.append(Product.level0 == level0)
    if level1:
        filters.append(Product.level1 == level1)
    if level2:
        filters.append(Product.level2 == level2)
    
    models = db.query(Product.level_2.distinct()).filter(
        and_(*filters)
    ).all()
    
    return [model[0] for model in models if model[0]]

@app.get("/hierarchy/skus")
async def get_skus_with_info(
    brand: Optional[str] = None,
    model: Optional[str] = None,
    level0: Optional[str] = None,
    level1: Optional[str] = None,
    level2: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Получить SKU с детальной информацией"""
    
    filters = [Product.is_available == True]
    
    if brand:
        filters.append(Product.brand == brand)
    if model:
        filters.append(Product.level_2 == model)
    if level0:
        filters.append(Product.level_0 == level0)
    if level1:
        filters.append(Product.level_1 == level1)
    if level2:
        filters.append(Product.level_2 == level2)
    
    # Получаем SKU с ценами
    results = db.query(Product, CurrentPrice).join(
        CurrentPrice, Product.sku == CurrentPrice.sku,
        isouter=True
    ).filter(and_(*filters)).all()
    
    skus_info = []
    for product, price in results:
        sku_data = {
            "sku": product.sku,
            "name": product.name,
            "brand": product.brand,
            "model": product.level_2 or "",
            "level0": product.level_0 or "",
            "level1": product.level_1 or "",
            "level2": product.level_2 or "",
            "price": price.price if price else 0.0,
            "currency": price.currency if price else "RUB",
            "stock": product.stock
        }
        skus_info.append(sku_data)
    
    return skus_info

@app.get("/debug/db-status")
async def debug_db_status(db: Session = Depends(get_db)):
    """Debug endpoint to check database status"""
    try:
        product_count = db.query(Product).count()
        category_count = db.query(Category).count()
        price_count = db.query(CurrentPrice).count()
        
        # Get sample products
        sample_products = db.query(Product).limit(5).all()
        
        return {
            "product_count": product_count,
            "category_count": category_count, 
            "price_count": price_count,
            "sample_products": [
                {
                    "id": p.id,
                    "name": p.name,
                    "model": p.level_2,
                    "sku": p.sku
                } for p in sample_products
            ],
            "database_url": Config.DATABASE_URL,
            "status": "ok"
        }
    except Exception as e:
        return {
            "error": str(e),
            "status": "error"
        }

@app.post("/import-single-product")
async def import_single_product(product_data: dict, db: Session = Depends(get_db)):
    """Добавить один товар через API"""
    try:
        # Валидация обязательных полей
        required_fields = ['name', 'brand', 'model']
        for field in required_fields:
            if not product_data.get(field):
                raise HTTPException(status_code=400, detail=f"Поле '{field}' обязательно для заполнения")
        
        # Генерируем SKU если не указан
        sku = product_data.get('sku')
        if not sku:
            brand_prefix = product_data['brand'][:3].upper()
            sku = f"{brand_prefix}{int(datetime.utcnow().timestamp())}"
        
        # Проверяем что SKU уникален
        existing_product = db.query(Product).filter(Product.sku == sku).first()
        if existing_product:
            raise HTTPException(status_code=400, detail=f"Товар с SKU '{sku}' уже существует")
        
        # Обрабатываем изображения
        images_data = product_data.get('images', [])
        if isinstance(images_data, list) and images_data:
            # Сохраняем как JSON массив
            images_json = json.dumps(images_data)
        else:
            images_json = None
        
        # Создаем новый товар
        specs = dict(product_data.get('specifications') or {})
        if product_data.get('color'):
            specs['color'] = product_data['color']
        if product_data.get('ram'):
            specs['ram'] = product_data['ram']
        if product_data.get('disk'):
            specs['disk'] = product_data['disk']
        if product_data.get('sim_config'):
            specs['sim_config'] = product_data['sim_config']

        new_product = Product(
            name=product_data['name'],
            sku=sku,
            brand=product_data['brand'],
            stock=product_data.get('stock', 0),
            is_available=product_data.get('is_available', True),
            specifications=json.dumps(specs),
            level_0=product_data.get('level0', ''),
            level_1=product_data.get('level1', ''),
            level_2=product_data.get('level2', '')
        )
        
        db.add(new_product)
        db.commit()
        db.refresh(new_product)
        # Ensure categories exist
        ensure_category_exists(db, product_data.get('level0'), product_data.get('level1'), product_data.get('level2'))
        
        # Добавляем цену, если она указана
        if 'price' in product_data and product_data['price']:
            price_obj = CurrentPrice(
                product_id=new_product.id,
                price=float(product_data['price']),
                old_price=float(product_data.get('old_price', product_data['price'])),
                currency='RUB',
                discount_percentage=0.0
            )
            db.add(price_obj)
            db.commit()
        
        # Создать запись изображений в ProductImage если есть изображения
        if images_json and product_data.get('level_2') and product_data.get('color'):
            product_image = ProductImage(
                level_2=product_data['level_2'],
                color=product_data['color'],
                img_list=images_json
            )
            db.add(product_image)
            db.commit()
        
        return {
            "success": True,
            "message": f"Товар '{product_data['name']}' успешно добавлен",
            "product_id": new_product.id,
            "sku": new_product.sku
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка добавления товара: {str(e)}")

@app.get("/export-products")
async def export_all_products(db: Session = Depends(get_db)):
    """Скачать полный ассортимент в Excel с всеми столбцами"""
    try:
        # Получить все товары с ценами
        results = db.query(Product, CurrentPrice).outerjoin(
            CurrentPrice, Product.sku == CurrentPrice.sku
        ).all()
        
        products_data = []
        for product, price in results:
            try:
                specifications = json.loads(product.specifications) if product.specifications else {}
            except json.JSONDecodeError:
                specifications = {}
            
            # Получить массив изображений
            images = get_product_images(product, db)
            
            products_data.append({
                'ID': product.id,
                'SKU': product.sku,
                'Название': product.name,
                'Описание': '',  # поле description удалено
                'Бренд': product.brand,
                'Категория': product.level_0 or '',
                'Уровень 0': product.level_0 or '',
                'Уровень 1': product.level_1 or '',
                'Уровень 2': product.level_2 or '',
                'Цвет': product.color or '',
                'Память': product.disk or '',
                'SIM': product.sim_config or '',
                'Цена': price.price if price else 0.0,
                'Старая цена': price.old_price if price else (price.price if price else 0.0),
                'Валюта': price.currency if price else 'RUB',
                'Скидка %': price.discount_percentage if price else 0.0,
                'Склад': product.stock,
                'В наличии': 'Да' if product.is_available else 'Нет',
                'Изображения': ' | '.join(images) if images else '',
                'Кол-во изображений': len(images),
                'Создано': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '',
                'Обновлено': product.updated_at.strftime('%Y-%m-%d %H:%M:%S') if product.updated_at else ''
            })
        
        # Создать DataFrame и Excel файл
        from io import BytesIO
        
        df = pd.DataFrame(products_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Ассортимент')
            
            # Стилизация
            workbook = writer.book
            worksheet = writer.sheets['Ассортимент']
            
            # Стили для заголовков
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        output.seek(0)
        
        return StreamingResponse(
            BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=assortment_full.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка экспорта: {str(e)}")

@app.get("/export-prices")
async def export_all_prices(db: Session = Depends(get_db)):
    """Скачать все цены в Excel"""
    try:
        # Получить все товары с ценами
        results = db.query(Product, CurrentPrice).outerjoin(
            CurrentPrice, Product.sku == CurrentPrice.sku
        ).filter(Product.is_available == True).all()
        
        prices_data = []
        for product, price in results:
            if price:  # Только товары с ценами
                prices_data.append({
                    'SKU': product.sku,
                    'Название товара': product.name,
                    'Бренд': product.brand,
                    'Текущая цена': price.price,
                    'Старая цена': price.old_price,
                    'Валюта': price.currency,
                    'Скидка %': f"{price.discount_percentage:.1f}%" if price.discount_percentage else "0%",
                    'Разница': f"{price.old_price - price.price:.0f}" if price.old_price and price.old_price > price.price else "0",
                    'Категория': product.level_0 or 'Без категории',
                    'В наличии': product.stock,
                    'Обновлено': price.last_updated.strftime('%Y-%m-%d %H:%M:%S') if hasattr(price, 'last_updated') and price.last_updated else 'Неизвестно'
                })
        
        if not prices_data:
            raise HTTPException(status_code=400, detail="Цены не найдены")
        
        # Создать DataFrame и Excel файл
        from io import BytesIO
        
        df = pd.DataFrame(prices_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Цены')
            
            # Стилизация
            workbook = writer.book
            worksheet = writer.sheets['Цены']
            
            # Стили для заголовков
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        output.seek(0)
        
        return StreamingResponse(
            BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=prices_full.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка экспорта цен: {str(e)}")

@app.get("/admin/schemes")
async def get_all_schemes(db: Session = Depends(get_db)):
    """Получить все схемы цветов и вариантов для админки"""
    from models import ModelColorScheme, ModelVariantScheme
    
    color_schemes = db.query(ModelColorScheme).all()
    variant_schemes = db.query(ModelVariantScheme).all()
    
    schemes_data = []
    for color_scheme in color_schemes:
        try:
            colors = json.loads(color_scheme.colors_json)
        except json.JSONDecodeError:
            colors = []
        
        # Найдем соответствующую схему вариантов
        variant_scheme = next((v for v in variant_schemes if v.model_key == color_scheme.model_key), None)
        variants = None
        if variant_scheme:
            try:
                variants = json.loads(variant_scheme.variants_json)
            except json.JSONDecodeError:
                variants = {}
        
        schemes_data.append({
            "model_key": color_scheme.model_key,
            "model_name": color_scheme.model_name,
            "brand": color_scheme.brand,
            "colors": colors,
            "variants": variants,
            "created_at": color_scheme.created_at.isoformat() if color_scheme.created_at else None
        })
    
    return {
        "schemes": schemes_data,
        "total": len(schemes_data)
    }

@app.post("/admin/schemes/{model_key}/colors")
async def update_color_scheme(model_key: str, colors_data: dict, db: Session = Depends(get_db)):
    """Обновить цветовую схему для модели"""
    from models import ModelColorScheme
    
    # Проверяем что colors_data содержит список цветов
    if "colors" not in colors_data:
        raise HTTPException(status_code=400, detail="Отсутствует поле 'colors'")
    
    colors = colors_data["colors"]
    if not isinstance(colors, list):
        raise HTTPException(status_code=400, detail="Поле 'colors' должно быть списком")
    
    # Ищем существующую схему
    color_scheme = db.query(ModelColorScheme).filter(
        ModelColorScheme.model_key == model_key
    ).first()
    
    if color_scheme:
        # Обновляем существующую схему
        color_scheme.colors_json = json.dumps(colors)
        color_scheme.updated_at = datetime.utcnow()
    else:
        # Создаем новую схему
        color_scheme = ModelColorScheme(
            model_key=model_key,
            model_name=colors_data.get("model_name", model_key),
            brand=colors_data.get("brand", "Unknown"),
            colors_json=json.dumps(colors)
        )
        db.add(color_scheme)
    
    db.commit()
    
    return {
        "success": True,
        "message": f"Цветовая схема для {model_key} обновлена",
        "colors_count": len(colors)
    }

@app.post("/admin/schemes/{model_key}/variants")
async def update_variant_scheme(model_key: str, variants_data: dict, db: Session = Depends(get_db)):
    """Обновить схему вариантов для модели"""
    from models import ModelVariantScheme
    
    # Проверяем что variants_data содержит варианты
    if "variants" not in variants_data:
        raise HTTPException(status_code=400, detail="Отсутствует поле 'variants'")
    
    variants = variants_data["variants"]
    if not isinstance(variants, dict):
        raise HTTPException(status_code=400, detail="Поле 'variants' должно быть объектом")
    
    # Ищем существующую схему
    variant_scheme = db.query(ModelVariantScheme).filter(
        ModelVariantScheme.model_key == model_key
    ).first()
    
    if variant_scheme:
        # Обновляем существующую схему
        variant_scheme.variants_json = json.dumps(variants)
        variant_scheme.updated_at = datetime.utcnow()
    else:
        # Создаем новую схему
        variant_scheme = ModelVariantScheme(
            model_key=model_key,
            model_name=variants_data.get("model_name", model_key),
            brand=variants_data.get("brand", "Unknown"),
            variants_json=json.dumps(variants)
        )
        db.add(variant_scheme)
    
    db.commit()
    
    return {
        "success": True,
        "message": f"Схема вариантов для {model_key} обновлена",
        "variants": list(variants.keys())
    }

@app.put("/products/{product_id}")
async def update_product(product_id: int, product_data: dict, db: Session = Depends(get_db)):
    """Обновить товар по ID"""
    try:
        # Найти товар
        product = db.query(Product).filter(Product.id == product_id).first()
        
        if not product:
            raise HTTPException(status_code=404, detail=f"Товар с ID {product_id} не найден")
        
        # Обновляем только переданные поля
        if 'name' in product_data:
            product.name = product_data['name']
        # if 'description' in product_data:
        #     product.description = product_data['description']  # поле удалено
        if 'brand' in product_data:
            product.brand = product_data['brand']
        if 'level_0' in product_data:
            product.level_0 = product_data['level_0']
        if 'level_1' in product_data:
            product.level_1 = product_data['level_1']
        if 'level_2' in product_data:
            product.level_2 = product_data['level_2']
        if 'color' in product_data:
            product.color = product_data['color']
        if 'disk' in product_data:
            product.disk = product_data['disk']
        if 'sim_config' in product_data:
            product.sim_config = product_data['sim_config']
        if 'is_available' in product_data:
            product.is_available = product_data['is_available']
        if 'stock' in product_data:
            product.stock = product_data['stock']
        
        # Обновляем изображения если указаны (в таблице ProductImage)
        if 'img_list' in product_data and product.level_2 and product.color:
            img_list_json = product_data['img_list']
            
            # Ищем или создаем запись в ProductImage
            product_image = db.query(ProductImage).filter(
                ProductImage.level_2 == product.level_2,
                ProductImage.color == product.color
            ).first()
            
            if product_image:
                product_image.img_list = img_list_json
            else:
                product_image = ProductImage(
                    level_2=product.level_2,
                    color=product.color,
                    img_list=img_list_json
                )
                db.add(product_image)
        
        # Обновление цены, если передана
        if 'price' in product_data:
            current_price = db.query(CurrentPrice).filter(CurrentPrice.sku == product.sku).first()
            if current_price:
                current_price.price = product_data['price']
                if 'old_price' in product_data:
                    current_price.old_price = product_data['old_price']
                current_price.last_updated = datetime.now()
            else:
                # Создаем новую цену если не существует
                new_price = CurrentPrice(
                    sku=product.sku,
                    price=product_data['price'],
                    old_price=product_data.get('old_price', product_data['price']),
                    currency=product_data.get('currency', 'RUB')
                )
                db.add(new_price)
        
        db.commit()
        db.refresh(product)
        
        return {
            "success": True,
            "message": f"Товар '{product.name}' успешно обновлен",
            "product": {
                "id": product.id,
                "sku": product.sku,
                "name": product.name,
                "brand": product.brand,
                "model": product.level_2 or "",
                "level_0": product.level_0,
                "level_1": product.level_1,
                "level_2": product.level_2
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка обновления товара: {str(e)}")

@app.delete("/products/{product_id}")
async def delete_product(product_id: int, db: Session = Depends(get_db)):
    """Удалить товар по ID"""
    try:
        # Найти товар
        product = db.query(Product).filter(Product.id == product_id).first()
        
        if not product:
            raise HTTPException(status_code=404, detail=f"Товар с ID {product_id} не найден")
        
        product_name = product.name
        product_sku = product.sku
        
        # Удалить связанные цены
        db.query(CurrentPrice).filter(CurrentPrice.sku == product_sku).delete()
        
        # Удалить товар
        db.delete(product)
        db.commit()
        
        return {
            "success": True,
            "message": f"Товар '{product_name}' (SKU: {product_sku}) успешно удален",
            "deleted_product_id": product_id,
            "deleted_sku": product_sku
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка удаления товара: {str(e)}")

@app.delete("/products/by-sku/{sku}")
async def delete_product_by_sku(sku: str, db: Session = Depends(get_db)):
    """Удалить товар по SKU"""
    try:
        # Найти товар
        product = db.query(Product).filter(Product.sku == sku).first()
        
        if not product:
            raise HTTPException(status_code=404, detail=f"Товар с SKU '{sku}' не найден")
        
        product_name = product.name
        product_id = product.id
        
        # Удалить связанные цены
        db.query(CurrentPrice).filter(CurrentPrice.sku == sku).delete()
        
        # Удалить товар
        db.delete(product)
        db.commit()
        
        return {
            "success": True,
            "message": f"Товар '{product_name}' (SKU: {sku}) успешно удален",
            "deleted_product_id": product_id,
            "deleted_sku": sku
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Ошибка удаления товара: {str(e)}")

@app.get("/level2-descriptions/{level_2}")
async def get_level2_description(level_2: str, db: Session = Depends(get_db)):
    """Get description and specifications for a level_2 product"""
    # Нормализуем регистр и лишние пробелы, ищем без учета регистра
    from sqlalchemy import func
    normalized = (level_2 or "").strip()
    description = db.query(Level2Description).filter(func.lower(Level2Description.level_2) == normalized.lower()).first()
    
    if not description:
        raise HTTPException(status_code=404, detail="Description not found")
    
    # Parse details if it's a JSON string
    details = {}
    if description.details:
        try:
            details = json.loads(description.details) if isinstance(description.details, str) else description.details
        except json.JSONDecodeError:
            details = {}
    
    return {
        "level_2": description.level_2,
        "description": description.description,
        "details": details
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=Config.HOST, port=Config.PORT)
